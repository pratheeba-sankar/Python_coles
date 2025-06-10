from openpyxl import Workbook, load_workbook
import os

EXCEL_FILE = "output.xlsx"

def write_to_excel(scenario, login_result, total_value):
    if os.path.exists(EXCEL_FILE):
        workbook = load_workbook(EXCEL_FILE)
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Results"
        sheet.append(["Scenario", "Login Result", "Total Value"])

    sheet = workbook["Results"]
    sheet.append([scenario, login_result, total_value])
    workbook.save(EXCEL_FILE)